import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def read_student_list(file):
    with open(file, 'r') as f:
        students = f.read().splitlines()
    return students

# Class details
classes_taken_dates = ["06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024", "03/09/2024", "17/09/2024", "01/10/2024"]
classes_missed_dates = ["10/09/2024"]  # New missing dates
exams_dates = ["24/09/2024"]  # Exam dates
all_dates = classes_taken_dates + classes_missed_dates + exams_dates  # Combine for Excel output


def process_attendance(attendance_file, students, classes_taken_dates):
    attendance_df = pd.read_csv(attendance_file)
    attendance_df['Timestamp'] = pd.to_datetime(attendance_df['Timestamp'], format="%d/%m/%Y %H:%M:%S")

    # Initialize attendance status for all students
    attendance_status = {student: {date: 0 for date in classes_taken_dates} for student in students}

    class_start_hour = 18  # Start time in hours
    class_end_hour = 20    # End time in hours

    # Process attendance for each student and class date
    for student in students:
        student_attendance = attendance_df[attendance_df['Roll'] == student]

        for date in classes_taken_dates:
            lecture_date = datetime.strptime(date, '%d/%m/%Y')  # Use the correct format for date parsing

            attended = student_attendance[
                (student_attendance['Timestamp'].dt.date == lecture_date.date()) &
                (student_attendance['Timestamp'].dt.hour >= class_start_hour) &
                (student_attendance['Timestamp'].dt.hour < class_end_hour)  # Non-inclusive upper limit
            ]

            if len(attended) == 1:
                attendance_status[student][date] = 1  # Partial attendance
            elif len(attended) == 2:
                attendance_status[student][date] = 2  # Full attendance
            elif len(attended) > 2:
                attendance_status[student][date] = len(attended)  # Proxy count for unusual cases

    return attendance_status

def compute_additional_columns(attendance_status, classes_taken_dates):
    # Convert attendance_status to DataFrame for easier manipulation
    attendance_df = pd.DataFrame.from_dict(attendance_status, orient='index', columns=classes_taken_dates)

    # Compute "Total Attendance Marked"
    attendance_df["Total Attendance Marked"] = attendance_df.sum(axis=1)

    # Compute "Total Attendance allowed"
    attendance_df["Total Attendance allowed"] = 2 * len(classes_taken_dates)

    # Compute "Total count of dates" (dates where student attended at least 1 class)
    attendance_df["Total count of dates"] = attendance_df[classes_taken_dates].apply(lambda row: (row >= 1).sum(), axis=1)

    attendance_df["Total correct count"] = attendance_df[classes_taken_dates].apply(
        lambda row: sum(min(att, 2) for att in row), axis=1
    )
    # Compute "Proxy"

    attendance_df["Proxy"] = attendance_df.apply(
        lambda row: row["Total Attendance Marked"] - row["Total correct count"] if row["Total Attendance Marked"] >  row["Total correct count"] else 0,
        axis=1
    )

    return attendance_df

def generate_excel_output(attendance_df, classes_taken_dates, output_file):
    wb = Workbook()
    ws = wb.active
    ws.append(['Student'] + classes_taken_dates + ["Total Attendance Marked", "Total Attendance allowed", "Total count of dates","Total correct count","Proxy"])

    # Define color fills for attendance status
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Write the attendance data to the Excel sheet
    for student, row in attendance_df.iterrows():
        ws.append([student] + row[classes_taken_dates].tolist() + [row["Total Attendance Marked"], row["Total Attendance allowed"], row["Total count of dates"],row["Total correct count"], row["Proxy"]])

    # Apply color coding to attendance cells
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=len(attendance_df) + 1, max_col=len(classes_taken_dates) + 1):
        for cell in row:
            if cell.value > 2:
                cell.fill = red_fill
            elif cell.value == 1:
                cell.fill = yellow_fill
            elif cell.value == 2:
                cell.fill = green_fill

    # Save the Excel file
    wb.save(output_file)

def main():
    student_list_file = 'stud_list.txt'
    attendance_file = 'input_attendance.csv'
    output_file = 'output_excel.xlsx'

    # Read the student list
    students = read_student_list(student_list_file)

    # Process attendance for the students
    attendance_status = process_attendance(attendance_file, students, classes_taken_dates)

    # Compute additional columns and convert to DataFrame
    attendance_df = compute_additional_columns(attendance_status, classes_taken_dates)

    # Generate the Excel output with the new columns
    generate_excel_output(attendance_df, classes_taken_dates, output_file)

if __name__ == '__main__':
    main()